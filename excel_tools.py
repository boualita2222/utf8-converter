#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Tools - Lire, Creer, Modifier des .xlsx
Python 3.14.2 + openpyxl
"""

import os
import sys

try:
    from colorama import init, Fore, Style
    init(autoreset=False, strip=False, convert=True)
except ImportError:
    class _D:
        def __getattr__(self, _): return ""
    Fore = Style = _D()

G  = Fore.GREEN  + Style.BRIGHT
Y  = Fore.YELLOW + Style.BRIGHT
C  = Fore.CYAN   + Style.BRIGHT
W  = Fore.WHITE
D  = Style.DIM   + Fore.WHITE
R  = Style.RESET_ALL
RE = Fore.RED    + Style.BRIGHT
LN = G + "=" * 55 + R

sys.path.insert(0, r"C:\PythonProjets")


# ════════════════════════════════════════════
#  MODULE 1 - CREER UN EXCEL SIMPLE
# ════════════════════════════════════════════
def creer_excel_simple(chemin):
    """Cree un fichier Excel simple"""
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill,
            Alignment, Border, Side)
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "UTF-8 Analyses"

        # ── Styles ───────────────────────
        # Couleurs
        vert_fonce  = "1e5c1e"
        vert_clair  = "e8f5e9"
        jaune       = "fff9c4"
        blanc       = "ffffff"
        gris        = "f5f5f5"

        # Police titre
        font_titre = Font(
            name="Calibri",
            bold=True,
            size=14,
            color="ffffff")

        # Police entete
        font_entete = Font(
            name="Calibri",
            bold=True,
            size=11,
            color="ffffff")

        # Police normale
        font_normal = Font(
            name="Calibri",
            size=11)

        # Fond vert fonce (entetes)
        fill_vert = PatternFill(
            start_color=vert_fonce,
            end_color=vert_fonce,
            fill_type="solid")

        # Fond vert clair
        fill_clair = PatternFill(
            start_color=vert_clair,
            end_color=vert_clair,
            fill_type="solid")

        # Fond jaune
        fill_jaune = PatternFill(
            start_color=jaune,
            end_color=jaune,
            fill_type="solid")

        # Fond gris
        fill_gris = PatternFill(
            start_color=gris,
            end_color=gris,
            fill_type="solid")

        # Bordure
        bordure_fine = Border(
            left=Side(style='thin',
                      color='cccccc'),
            right=Side(style='thin',
                       color='cccccc'),
            top=Side(style='thin',
                     color='cccccc'),
            bottom=Side(style='thin',
                        color='cccccc'))

        # Alignement centre
        align_centre = Alignment(
            horizontal='center',
            vertical='center')

        # Alignement gauche
        align_gauche = Alignment(
            horizontal='left',
            vertical='center')

        # ── TITRE ────────────────────────
        ws.merge_cells('A1:F1')
        titre_cell = ws['A1']
        titre_cell.value = "UTF-8 Converter Pro - Rapport Excel"
        titre_cell.font  = Font(
            name="Calibri", bold=True,
            size=16, color="ffffff")
        titre_cell.fill      = fill_vert
        titre_cell.alignment = align_centre
        ws.row_dimensions[1].height = 35

        # ── SOUS-TITRE ───────────────────
        ws.merge_cells('A2:F2')
        from datetime import datetime
        date_str = datetime.now().strftime(
            "%d/%m/%Y %H:%M")
        sous_titre = ws['A2']
        sous_titre.value = "Genere le : " + date_str
        sous_titre.font  = Font(
            name="Calibri", size=10,
            color="666666", italic=True)
        sous_titre.alignment = align_centre
        ws.row_dimensions[2].height = 20

        # ── EN-TETES ─────────────────────
        entetes = [
            'Texte', 'Hex UTF-8',
            'Nb Chars', 'Nb Octets',
            'Octets/Char', 'Type'
        ]
        for col, entete in enumerate(
                entetes, 1):
            cell = ws.cell(
                row=3, column=col,
                value=entete)
            cell.font      = font_entete
            cell.fill      = fill_vert
            cell.alignment = align_centre
            cell.border    = bordure_fine
        ws.row_dimensions[3].height = 25

        # ── DONNEES ──────────────────────
        try:
            from utf8_package.encodeur import \
                encoder_texte

            textes = [
                "cafe",
                "Bonjour",
                "Python",
                "Django",
                "UTF-8",
                "Alger",
            ]

            for i, texte in enumerate(
                    textes, 4):
                res = encoder_texte(texte)
                nb_chars  = res['nb_chars']
                nb_octets = res['nb_octets']
                ratio     = round(
                    nb_octets / nb_chars, 2)

                # Type de texte
                if ratio == 1.0:
                    type_txt = "ASCII pur"
                elif ratio < 2.0:
                    type_txt = "Latin"
                elif ratio < 3.0:
                    type_txt = "Symboles"
                else:
                    type_txt = "Emojis"

                # Alternance couleurs
                fill_row = (fill_clair
                            if i % 2 == 0
                            else PatternFill(
                    start_color=blanc,
                    end_color=blanc,
                    fill_type="solid"))

                donnees = [
                    texte,
                    res['hex'][:30],
                    nb_chars,
                    nb_octets,
                    ratio,
                    type_txt,
                ]

                for col, val in enumerate(
                        donnees, 1):
                    cell = ws.cell(
                        row=i,
                        column=col,
                        value=val)
                    cell.font      = font_normal
                    cell.fill      = fill_row
                    cell.border    = bordure_fine
                    cell.alignment = (
                        align_centre
                        if col > 2
                        else align_gauche)

                ws.row_dimensions[i].height = 20

        except Exception as e:
            print(D + "  Note UTF-8 : " +
                  str(e) + R)
            # Donnees de demo
            demo = [
                ["cafe", "63 61 66 65",
                 4, 4, 1.0, "ASCII pur"],
                ["Bonjour",
                 "42 6F 6E 6A 6F 75 72",
                 7, 7, 1.0, "ASCII pur"],
                ["Python",
                 "50 79 74 68 6F 6E",
                 6, 6, 1.0, "ASCII pur"],
            ]
            for i, row in enumerate(demo, 4):
                for col, val in enumerate(
                        row, 1):
                    cell = ws.cell(
                        row=i,
                        column=col,
                        value=val)
                    cell.font   = font_normal
                    cell.border = bordure_fine

        # ── LARGEURS COLONNES ────────────
        largeurs = {
            'A': 15,
            'B': 35,
            'C': 12,
            'D': 12,
            'E': 14,
            'F': 15,
        }
        for col, largeur in largeurs.items():
            ws.column_dimensions[col].width = \
                largeur

        # ── FIGER LES VOLETS ─────────────
        ws.freeze_panes = 'A4'

        # ── FILTRE AUTO ──────────────────
        ws.auto_filter.ref = "A3:F3"

        wb.save(chemin)
        print(G + "\n  Excel cree : " +
              chemin + R)
        return True

    except Exception as e:
        print(RE + "  Erreur : " + str(e) + R)
        return False


# ════════════════════════════════════════════
#  MODULE 2 - CREER RAPPORT AVEC GRAPHIQUE
# ════════════════════════════════════════════
def creer_excel_graphique(chemin):
    """Cree un Excel avec graphique"""
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment)
        from openpyxl.chart import BarChart, Reference

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stats UTF-8"

        # En-tetes
        ws['A1'] = 'Type'
        ws['B1'] = 'Nb Chars'
        ws['C1'] = 'Nb Octets'

        font_h = Font(bold=True, color="ffffff")
        fill_h = PatternFill(
            start_color="1e5c1e",
            end_color="1e5c1e",
            fill_type="solid")

        for col in ['A', 'B', 'C']:
            ws[col + '1'].font = font_h
            ws[col + '1'].fill = fill_h

        # Donnees
        donnees = [
            ("ASCII (1 oct)",   20, 20),
            ("Latin (2 oct)",    5, 10),
            ("Symboles (3 oct)", 2,  6),
            ("Emojis (4 oct)",   1,  4),
        ]

        for i, (type_txt, chars, octets) \
                in enumerate(donnees, 2):
            ws.cell(row=i, column=1,
                    value=type_txt)
            ws.cell(row=i, column=2,
                    value=chars)
            ws.cell(row=i, column=3,
                    value=octets)

        # Largeurs
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12

        # Graphique barre
        chart = BarChart()
        chart.type    = "col"
        chart.title   = "Distribution UTF-8"
        chart.y_axis.title = "Nombre"
        chart.x_axis.title = "Type"
        chart.style   = 10
        chart.width   = 15
        chart.height  = 10

        data = Reference(ws,
            min_col=2, max_col=3,
            min_row=1, max_row=5)
        cats = Reference(ws,
            min_col=1,
            min_row=2, max_row=5)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E1")

        wb.save(chemin)
        print(G + "\n  Excel graphique : " +
              chemin + R)
        return True

    except Exception as e:
        print(RE + "  Erreur : " + str(e) + R)
        return False


# ════════════════════════════════════════════
#  MODULE 3 - LIRE UN EXCEL
# ════════════════════════════════════════════
def lire_excel(chemin):
    """Lit et affiche le contenu d'un Excel"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(chemin)

        print(G + "\n  Excel : " + chemin + R)
        print(C + "  Feuilles : " +
              str(wb.sheetnames) + R)

        for nom_feuille in wb.sheetnames:
            ws = wb[nom_feuille]
            print(Y + "\n  Feuille : " +
                  nom_feuille + R)
            print(D + "  Lignes : " +
                  str(ws.max_row) +
                  "  Colonnes : " +
                  str(ws.max_column) + R)
            print()

            # Afficher les 10 premieres lignes
            for i, row in enumerate(
                    ws.iter_rows(
                        max_row=10,
                        values_only=True), 1):
                if any(v is not None
                       for v in row):
                    ligne = "  "
                    for val in row:
                        if val is not None:
                            ligne += str(
                                val)[:15].ljust(16)
                    print(W + ligne + R)

        return True

    except FileNotFoundError:
        print(RE + "  Fichier introuvable : " +
              chemin + R)
        return False
    except Exception as e:
        print(RE + "  Erreur : " + str(e) + R)
        return False


# ════════════════════════════════════════════
#  MODULE 4 - MODIFIER UN EXCEL
# ════════════════════════════════════════════
def modifier_excel(chemin):
    """Ajoute une feuille a un Excel existant"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.load_workbook(chemin)

        # Nouvelle feuille
        ws2 = wb.create_sheet("Resume")
        ws2['A1'] = "Resume des analyses"
        ws2['A1'].font = Font(
            bold=True, size=14,
            color="1e5c1e")

        ws2['A3'] = "Total analyses :"
        ws2['B3'] = 6

        ws2['A4'] = "Octets ASCII :"
        ws2['B4'] = 40

        ws2['A5'] = "Octets non-ASCII :"
        ws2['B5'] = 0

        ws2['A6'] = "Ratio moyen :"
        ws2['B6'] = 1.0

        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 10

        wb.save(chemin)
        print(G + "\n  Excel modifie : " +
              chemin + R)
        return True

    except Exception as e:
        print(RE + "  Erreur : " + str(e) + R)
        return False


# ════════════════════════════════════════════
#  PROGRAMME PRINCIPAL
# ════════════════════════════════════════════
def main():
    os.system("cls" if os.name == "nt"
              else "clear")
    print(G + """
╔══════════════════════════════════════════════════════╗
║         EXCEL TOOLS - Python 3.14.2                 ║
║         Lire, Creer, Modifier des .xlsx             ║
╚══════════════════════════════════════════════════════╝""" + R)

    chemin1 = r"C:\PythonProjets\rapport_utf8.xlsx"
    chemin2 = r"C:\PythonProjets\stats_utf8.xlsx"

    # Test 1 - Creer Excel avec mise en forme
    print(Y + "\n  [1] Creation Excel professionnel...\n" + R)
    creer_excel_simple(chemin1)

    # Test 2 - Creer Excel avec graphique
    print(Y + "\n  [2] Creation Excel avec graphique...\n" + R)
    creer_excel_graphique(chemin2)

    # Test 3 - Lire l'Excel cree
    print(Y + "\n  [3] Lecture Excel...\n" + R)
    lire_excel(chemin1)

    # Test 4 - Modifier l'Excel
    print(Y + "\n  [4] Modification Excel...\n" + R)
    modifier_excel(chemin1)

    print()
    print(LN)
    print(G + "\n  Excel Tools OK !\n" + R)
    print(W + "  Fichiers crees :\n"
          "  C:\\PythonProjets\\rapport_utf8.xlsx\n"
          "  C:\\PythonProjets\\stats_utf8.xlsx\n" + R)
    input(D + "  Entree pour quitter..." + R)


if __name__ == "__main__":
    main()